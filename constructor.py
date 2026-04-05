"""Constructor module: writes the new working paper sheet from the prior year template."""
from __future__ import annotations

import re
import zipfile as _zf
from copy import copy as copy_obj
from dataclasses import dataclass
from typing import Literal, Any
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Font, PatternFill, Side

from processor import WorkbookData, TrialBalanceRow


# ── Legacy dataclass kept for deterministic tests ────────────────────────────

@dataclass
class WorkingPaperRow:
    row_type: Literal["section_header", "account", "group_total"]
    group: str | None
    account_number: str | None
    details: str | None
    debit: float | None
    credit: float | None
    opening_balance: float | None
    notes: str | None


# ── Style constants ───────────────────────────────────────────────────────────

EXISTING_FILL = PatternFill(fill_type="solid", fgColor="FF92D050")  # green
MISSING_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF00")  # yellow
NEW_FILL      = PatternFill(fill_type="solid", fgColor="FF00B0F0")  # blue

GREEN_FILL  = EXISTING_FILL
YELLOW_FILL = MISSING_FILL

_THIN_SIDE   = Side(style='thin')
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_XMLNS_R = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_account(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    if s and s.isdigit():
        return str(int(s))
    return s or None


def _is_account_cell(val: Any) -> bool:
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip()
        return bool(s) and s.isdigit()
    return False


def _split_debit_credit(
    debit: float | None, credit: float | None
) -> tuple[float | None, float | None]:
    d = debit if (debit is not None and debit > 0) else None
    c = credit if (credit is not None and credit > 0) else None
    if d is not None and c is not None:
        net = d - c
        return (net, None) if net >= 0 else (None, -net)
    if d is not None:
        return d, None
    if c is not None:
        return None, c
    return 0.0, None


def _remap_row_refs(formula: str, orig_to_new: dict[int, int]) -> str:
    def replacer(m: re.Match) -> str:
        col = m.group(1)
        row = int(m.group(2))
        return f"{col}{orig_to_new.get(row, row)}"
    return re.sub(r'([A-Za-z]+)(\d+)', replacer, formula)


def _parse_col_sum_range(formula: str) -> tuple[str, int, int] | None:
    m = re.match(r"=SUM\(([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)\)", formula or "", re.IGNORECASE)
    if m and m.group(1).upper() == m.group(3).upper():
        return (m.group(1).upper(), int(m.group(2)), int(m.group(4)))
    return None


def _extend_sum_formula(
    formula: str,
    orig_to_new: dict[int, int],
    new_account_insertions: list[tuple[int, int | None]],
    last_template_acct_orig: int,
) -> str:
    orig_range = _parse_col_sum_range(formula)
    remapped = _remap_row_refs(formula, orig_to_new)
    if orig_range is None:
        return remapped
    col_letter, orig_start, orig_end = orig_range
    parsed_new = _parse_col_sum_range(remapped)
    if parsed_new is None:
        return remapped
    _, new_start, new_end = parsed_new
    for acct_row, after_orig in new_account_insertions:
        if after_orig is not None:
            should_extend = orig_start <= after_orig <= orig_end
        else:
            should_extend = orig_end >= last_template_acct_orig
        if should_extend and acct_row > new_end:
            new_end = acct_row
    return f"=SUM({col_letter}{new_start}:{col_letter}{new_end})"


# ── F / G column neutral formatting ──────────────────────────────────────────

_COLOR_TAG_RE = re.compile(
    r'\[(?:Red|Blue|Green|Yellow|Black|White|Cyan|Magenta|Color\s*\d+)\]',
    re.IGNORECASE,
)


def _strip_bracket_neg(fmt: str) -> str:
    if not fmt:
        return '#,##0;-#,##0'
    fmt = _COLOR_TAG_RE.sub('', fmt)
    fmt = re.sub(r'\*.', '', fmt)
    parts = fmt.split(';')
    if len(parts) >= 2:
        parts[1] = re.sub(r'\(([^)]+)\)', r'-\1', parts[1])
    return ';'.join(parts)


def _neutralize_cell(cell: Any) -> None:
    cell.number_format = _strip_bracket_neg(cell.number_format or '#,##0')
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        underline=f.underline, strike=f.strike, color='FF000000',
    )


# ── Border helper ─────────────────────────────────────────────────────────────

def _ensure_border(cell: Any) -> None:
    b = cell.border
    cell.border = Border(
        left=b.left     if (b.left   and b.left.style)   else _THIN_SIDE,
        right=b.right   if (b.right  and b.right.style)  else _THIN_SIDE,
        top=b.top       if (b.top    and b.top.style)    else _THIN_SIDE,
        bottom=b.bottom if (b.bottom and b.bottom.style) else _THIN_SIDE,
    )


# ── Cell copy helper ──────────────────────────────────────────────────────────

def _copy_cell(src: Any, dst: Any) -> None:
    dst.value = src.value
    if src.has_style:
        dst.font = copy_obj(src.font)
        dst.fill = copy_obj(src.fill)
        dst.border = copy_obj(src.border)
        dst.alignment = copy_obj(src.alignment)
        dst.number_format = src.number_format


# ── Cross-sheet formula remapping ─────────────────────────────────────────────

def _remap_cross_sheet_refs(
    wb: Workbook,
    new_sheet_name: str,
    orig_to_new: dict[int, int],
) -> None:
    if not orig_to_new:
        return
    escaped_name = re.escape(new_sheet_name)
    pattern = re.compile(
        r"('" + escaped_name + r"'!|" + escaped_name + r"!)([A-Z]+)(\d+)",
        re.IGNORECASE,
    )
    def _remap_match(m: re.Match) -> str:
        return f"{m.group(1)}{m.group(2)}{orig_to_new.get(int(m.group(3)), int(m.group(3)))}"

    for ws_name in wb.sheetnames:
        if ws_name == new_sheet_name:
            continue
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                if "#REF!" in cell.value or new_sheet_name not in cell.value:
                    continue
                new_formula = pattern.sub(_remap_match, cell.value)
                if new_formula != cell.value:
                    cell.value = new_formula


# ── Workbook parts preservation (drawings, printer settings) ─────────────────

def _preserve_workbook_parts(input_path: Path, output_path: Path) -> None:
    """Restore drawings and other content that openpyxl silently drops on save.

    openpyxl discards <drawing r:id=...> elements, worksheet .rels files,
    drawing XML, and printer-settings binaries.  Without them Excel shows a
    "We found a problem" repair dialog.

    Steps:
    1. Copy worksheet .rels files (with updated sheet numbers) from input.
    2. Copy every resource they reference (drawings, printer settings).
    3. Re-inject <drawing>/<legacyDrawing> elements into the sheet XML,
       adding the required xmlns:r namespace declaration.
    4. Update Content_Types.xml to declare the restored resource types.
    """

    def _attr(xml: str, name: str) -> str | None:
        """Extract attribute value from an XML element string."""
        m = re.search(rf'\b{re.escape(name)}="([^"]*)"', xml)
        return m.group(1) if m else None

    def _sheet_file_map(path: Path) -> dict[str, str]:
        """Return {sheet_name: xml_filename} for every sheet."""
        with _zf.ZipFile(path) as z:
            wb_xml   = z.read("xl/workbook.xml").decode()
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode()

        # sheet name -> r:id (handles both attribute orders)
        name_rid: dict[str, str] = {}
        for elem in re.findall(r'<sheet\b[^/]*/>', wb_xml):
            name = _attr(elem, 'name')
            rid  = _attr(elem, 'r:id')
            if name and rid:
                name_rid[name] = rid

        # r:id -> filename (handles both attribute orders)
        rid_file: dict[str, str] = {}
        for elem in re.findall(r'<Relationship\b[^/]*/>', rels_xml):
            rid    = _attr(elem, 'Id')
            target = _attr(elem, 'Target')
            if rid and target:
                rid_file[rid] = target.split('/')[-1]

        return {n: rid_file[r] for n, r in name_rid.items() if r in rid_file}

    in_map  = _sheet_file_map(input_path)
    out_map = _sheet_file_map(output_path)

    with _zf.ZipFile(input_path) as z_in:
        in_names = set(z_in.namelist())
        sheets_with_rels = {
            name: (in_map[name], f"xl/worksheets/_rels/{in_map[name]}.rels")
            for name in in_map
            if f"xl/worksheets/_rels/{in_map[name]}.rels" in in_names
        }

    if not sheets_with_rels:
        return

    # Read entire output zip into memory
    out_files: dict[str, bytes] = {}
    with _zf.ZipFile(output_path) as z:
        for name in z.namelist():
            out_files[name] = z.read(name)

    with _zf.ZipFile(input_path) as z_in:
        for sheet_name, (old_file, old_rels_path) in sheets_with_rels.items():
            new_file = out_map.get(sheet_name)
            if not new_file:
                continue

            new_rels_path = f"xl/worksheets/_rels/{new_file}.rels"
            rels_text = z_in.read(old_rels_path).decode()
            out_files[new_rels_path] = rels_text.encode()

            # Copy every resource referenced by the .rels file
            for target in re.findall(r'Target="([^"]+)"', rels_text):
                # Resolve relative path against the .rels file location
                parts = f"xl/worksheets/_rels/{old_file}.rels".split('/')
                for seg in target.split('/'):
                    if seg == '..':
                        parts.pop()
                    elif seg:
                        parts.append(seg)
                resource_path = '/'.join(parts)
                if resource_path in in_names:
                    out_files[resource_path] = z_in.read(resource_path)

            # Re-inject <drawing> / <legacyDrawing> elements openpyxl stripped
            old_sheet_xml = z_in.read(f"xl/worksheets/{old_file}").decode()
            new_sheet_xml = out_files.get(f"xl/worksheets/{new_file}", b"").decode()

            inject = ""
            for tag in ("drawing", "legacyDrawing"):
                m = re.search(rf'<{tag}[^/]*/>', old_sheet_xml)
                if m and m.group(0) not in new_sheet_xml:
                    inject += m.group(0)

            if inject:
                if _XMLNS_R not in new_sheet_xml:
                    new_sheet_xml = new_sheet_xml.replace(
                        "<worksheet ", f"<worksheet {_XMLNS_R} ", 1
                    )
                new_sheet_xml = new_sheet_xml.replace("</worksheet>", inject + "</worksheet>")
                out_files[f"xl/worksheets/{new_file}"] = new_sheet_xml.encode()

    # Update Content_Types.xml to declare new resource types
    ct = out_files.get("[Content_Types].xml", b"").decode()
    with _zf.ZipFile(input_path) as z_in:
        in_ct = z_in.read("[Content_Types].xml").decode()

    for m in re.finditer(r'(<Override[^>]+PartName="([^"]+)"[^>]*/>)', in_ct):
        override_el, part = m.group(1), m.group(2)
        if part.lstrip('/') in out_files and override_el not in ct:
            ct = ct.replace("</Types>", override_el + "</Types>")

    out_files["[Content_Types].xml"] = ct.encode()

    # Rewrite the output zip
    tmp = output_path.with_suffix(".tmp.xlsx")
    with _zf.ZipFile(tmp, "w", compression=_zf.ZIP_DEFLATED) as z_out:
        for name, data in out_files.items():
            z_out.writestr(name, data)
    tmp.replace(output_path)


# ── Core template-based generation ───────────────────────────────────────────

def _build_output_rows(
    prior_ws: Worksheet,
    new_accounts: list[TrialBalanceRow],
) -> list[dict]:
    new_by_group: dict[str, list[TrialBalanceRow]] = {}
    for tb in new_accounts:
        g = _normalize_account(tb.group) or ""
        new_by_group.setdefault(g, []).append(tb)

    template: list[dict] = []
    for row in prior_ws.iter_rows():
        cells = list(row)
        col_b = cells[1].value if len(cells) > 1 else None
        group_val = cells[0].value if cells else None
        is_account = _is_account_cell(col_b)
        template.append({
            "orig_row": cells[0].row,
            "cells": cells,
            "account_num": _normalize_account(col_b),
            "group": _normalize_account(group_val) if is_account else None,
            "is_account": is_account,
        })

    group_last_idx: dict[str, int] = {}
    for i, t in enumerate(template):
        if t["is_account"] and t["group"]:
            group_last_idx[t["group"]] = i

    output: list[dict] = []
    inserted_groups: set[str] = set()

    for i, t in enumerate(template):
        output.append(t)
        if t["is_account"] and t["group"] and group_last_idx.get(t["group"]) == i:
            group = t["group"]
            if group in new_by_group and group not in inserted_groups:
                for tb in sorted(
                    new_by_group[group],
                    key=lambda x: _normalize_account(x.account_number) or "",
                ):
                    output.append({"is_new": True, "tb": tb, "inserted_after_orig_row": t["orig_row"]})
                inserted_groups.add(group)

    remaining = sorted(
        [tb for tb in new_accounts if (_normalize_account(tb.group) or "") not in inserted_groups],
        key=lambda tb: (
            _normalize_account(tb.group) or "",
            _normalize_account(tb.account_number) or "",
        ),
    )
    if remaining:
        last_tmpl_acct_idx = max(
            (i for i, row in enumerate(output)
             if not row.get("is_new") and row["is_account"]),
            default=len(output) - 1,
        )
        last_tmpl_acct_orig = output[last_tmpl_acct_idx]["orig_row"]
        insert_at = last_tmpl_acct_idx + 1
        for tb in remaining:
            output.insert(insert_at, {
                "is_new": True,
                "tb": tb,
                "inserted_after_orig_row": last_tmpl_acct_orig,
            })
            insert_at += 1

    return output


def _write_output_rows(
    ws: Worksheet,
    output_rows: list[dict],
    tb_by_account: dict[str, TrialBalanceRow],
    prior_year_balances: dict[str, float],
    prior_ws: Worksheet,
) -> dict[int, int]:
    for col_letter, col_dim in prior_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width
        ws.column_dimensions[col_letter].hidden = col_dim.hidden

    orig_to_new: dict[int, int] = {}
    assigned_accounts: set[str] = set()
    new_account_insertions: list[tuple[int, int | None]] = []
    template_rows_written: list[tuple[int, dict]] = []

    for new_row_num, row in enumerate(output_rows, start=1):

        if row.get("is_new"):
            tb: TrialBalanceRow = row["tb"]
            group = _normalize_account(tb.group) or ""
            debit, credit = _split_debit_credit(tb.debit, tb.credit)
            acct_str = _normalize_account(tb.account_number)

            ws.cell(new_row_num, 1).value = int(group) if group.isdigit() else group
            ws.cell(new_row_num, 2).value = (
                int(acct_str) if (acct_str and acct_str.isdigit()) else acct_str
            )
            ws.cell(new_row_num, 3).value = tb.account_name
            ws.cell(new_row_num, 4).value = debit
            ws.cell(new_row_num, 5).value = credit
            ws.cell(new_row_num, 7).value = f"=D{new_row_num}-E{new_row_num}+F{new_row_num}"
            for col in range(1, 6):
                ws.cell(new_row_num, col).fill = copy_obj(NEW_FILL)

            new_account_insertions.append((new_row_num, row.get("inserted_after_orig_row")))
            continue

        orig_row: int = row["orig_row"]
        orig_to_new[orig_row] = new_row_num
        cells: list = row["cells"]

        for cell in cells:
            _copy_cell(cell, ws.cell(new_row_num, cell.column))

        src_height = prior_ws.row_dimensions[orig_row].height
        if src_height:
            ws.row_dimensions[new_row_num].height = src_height

        if row["is_account"]:
            account_num = row["account_num"] or ""

            tb_row = tb_by_account.get(account_num)
            if tb_row and account_num not in assigned_accounts:
                debit, credit = _split_debit_credit(tb_row.debit, tb_row.credit)
                ws.cell(new_row_num, 4).value = debit
                ws.cell(new_row_num, 5).value = credit
                assigned_accounts.add(account_num)
                for col in range(1, 6):
                    ws.cell(new_row_num, col).fill = copy_obj(EXISTING_FILL)
            elif tb_row:
                ws.cell(new_row_num, 4).value = None
                ws.cell(new_row_num, 5).value = None
                for col in range(1, 6):
                    ws.cell(new_row_num, col).fill = copy_obj(EXISTING_FILL)
            else:
                ws.cell(new_row_num, 4).value = None
                ws.cell(new_row_num, 5).value = None
                for col in range(1, 6):
                    ws.cell(new_row_num, col).fill = copy_obj(MISSING_FILL)

            ws.cell(new_row_num, 7).value = f"=D{new_row_num}-E{new_row_num}+F{new_row_num}"

        template_rows_written.append((new_row_num, row))

    last_template_acct_orig = max(
        (row["orig_row"] for _, row in template_rows_written if row["is_account"]),
        default=0,
    )

    for new_row_num, row in template_rows_written:
        h_cell = ws.cell(new_row_num, 8)
        if isinstance(h_cell.value, str) and h_cell.value.startswith("="):
            h_cell.value = _extend_sum_formula(
                h_cell.value, orig_to_new, new_account_insertions, last_template_acct_orig
            )

        if row["is_account"]:
            f_cell = ws.cell(new_row_num, 6)
            if isinstance(f_cell.value, str) and f_cell.value.startswith("="):
                f_cell.value = _remap_row_refs(f_cell.value, orig_to_new)
        else:
            for col_idx in range(1, 12):
                if col_idx == 8:
                    continue
                cell = ws.cell(new_row_num, col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _extend_sum_formula(
                        cell.value, orig_to_new, new_account_insertions, last_template_acct_orig
                    )

    total_rows = len(output_rows)
    for row_num in range(1, total_rows + 1):
        f_cell = ws.cell(row_num, 6)
        if f_cell.value is not None:
            _neutralize_cell(f_cell)
        g_cell = ws.cell(row_num, 7)
        if g_cell.value is not None:
            _neutralize_cell(g_cell)
        for col_num in range(1, 9):
            _ensure_border(ws.cell(row_num, col_num))

    for col_letter in ('F', 'G'):
        dim = ws.column_dimensions[col_letter]
        if not dim.width or dim.width < 16:
            dim.width = 16

    return orig_to_new


# ── Public entry point ────────────────────────────────────────────────────────

def generate_working_paper(data: WorkbookData, output_path: Path | None = None) -> Path:

    wb = openpyxl.load_workbook(data.workbook_path)
    prior_ws: Worksheet = wb[str(data.prior_year)]

    tb_by_account: dict[str, TrialBalanceRow] = {}
    for tb in data.trial_balance_rows:
        if tb.row_type == "account" and tb.account_number:
            key = _normalize_account(tb.account_number) or ""
            if key:
                tb_by_account[key] = tb

    prior_accounts: set[str] = set()
    for row in prior_ws.iter_rows():
        cells = list(row)
        col_b = cells[1].value if len(cells) > 1 else None
        if _is_account_cell(col_b):
            acct = _normalize_account(col_b)
            if acct:
                prior_accounts.add(acct)

    new_accounts = [
        tb
        for tb in data.trial_balance_rows
        if tb.row_type == "account"
        and tb.account_number
        and (_normalize_account(tb.account_number) or "") not in prior_accounts
    ]

    output_rows = _build_output_rows(prior_ws, new_accounts)

    sheet_name = str(data.new_year)
    if sheet_name in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' already exists")
    new_ws: Worksheet = wb.create_sheet(sheet_name)

    new_ws.sheet_view.rightToLeft = prior_ws.sheet_view.rightToLeft

    target_name = f"מאזן {data.new_year}"
    if target_name in wb.sheetnames:
        target_idx = wb.sheetnames.index(target_name)
        current_idx = wb.sheetnames.index(sheet_name)
        offset = (target_idx + 1) - current_idx
        if offset:
            wb.move_sheet(sheet_name, offset)

    orig_to_new = _write_output_rows(
        new_ws, output_rows, tb_by_account, data.prior_year_balances, prior_ws
    )

    _remap_cross_sheet_refs(wb, sheet_name, orig_to_new)

    for merged_range in prior_ws.merged_cells.ranges:
        try:
            new_ws.merge_cells(str(merged_range))
        except Exception:
            pass

    if output_path is None:
        output_dir = Path(__file__).parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / data.workbook_path.name
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(output_path)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save: '{output_path}' is open in another application. Close it and try again."
        )
    except IOError as e:
        raise RuntimeError(f"Failed to save workbook: {e}")

    # Restore drawings and other content dropped by openpyxl
    _preserve_workbook_parts(data.workbook_path, output_path)

    return output_path


# ── Legacy helpers (kept for deterministic tests) ─────────────────────────────

def determine_output_path(input_path: Path) -> Path:
    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / input_path.name


def build_sheet(ws: Worksheet, rows: list[WorkingPaperRow], title: str) -> None:
    ws.cell(row=1, column=1, value=title)
    headers = [
        (1, "קבוצה"), (2, "מס' כרטיס"), (3, "פרטים"),
        (4, "חובה"), (5, "זכות"), (6, "פ.נ"), (7, "יתרה"), (8, "הערות"),
    ]
    for col, header in headers:
        ws.cell(row=2, column=col, value=header)

    current_row = 3
    for row in rows:
        if row.row_type == "section_header":
            ws.cell(row=current_row, column=3, value=row.details or "")
        elif row.row_type == "account":
            ws.cell(row=current_row, column=1, value=row.group)
            ws.cell(row=current_row, column=2, value=row.account_number)
            ws.cell(row=current_row, column=3, value=row.details)
            if row.debit is not None:
                ws.cell(row=current_row, column=4, value=row.debit)
            if row.credit is not None:
                ws.cell(row=current_row, column=5, value=row.credit)
            ws.cell(row=current_row, column=6, value=row.opening_balance)
            ws.cell(row=current_row, column=7, value=f"=F{current_row}+D{current_row}-E{current_row}")
        elif row.row_type == "group_total":
            ws.cell(row=current_row, column=3, value=row.details)
            if row.debit is not None:
                ws.cell(row=current_row, column=4, value=row.debit)
            if row.credit is not None:
                ws.cell(row=current_row, column=5, value=row.credit)
        current_row += 1
