"""Integration tests for the full pipeline (no Claude CLI required)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from processor import load_workbook_data, WorkbookData
from constructor import WorkingPaperRow, generate_working_paper


def make_sample_workbook(tmp_path: Path, year: int = 2024) -> Path:
    """Create a minimal synthetic workbook for testing."""
    prior = year - 1
    wb = openpyxl.Workbook()

    # Remove default sheet
    del wb[wb.sheetnames[0]]

    # --- Prior trial balance sheet (minimal, just needs to exist) ---
    ws_tb_prior = wb.create_sheet(f"מאזן {prior}")
    ws_tb_prior.cell(1, 1, f"Company {prior}")
    ws_tb_prior.cell(2, 1, "מאזן בוחן")
    ws_tb_prior.cell(3, 1, f"לשנת {prior}")
    ws_tb_prior.cell(4, 2, "מיון")
    ws_tb_prior.cell(4, 3, "חשבון")
    ws_tb_prior.cell(4, 4, "שם חשבון")
    ws_tb_prior.cell(4, 5, "חובה")
    ws_tb_prior.cell(4, 6, "זכות")
    ws_tb_prior.cell(4, 7, "הפרש")

    # --- Prior working paper sheet ---
    ws_prior = wb.create_sheet(str(prior))
    ws_prior.cell(1, 1, str(prior))
    ws_prior.cell(2, 1, "קבוצה")
    ws_prior.cell(2, 2, "מס' כרטיס")
    ws_prior.cell(2, 3, "פרטים")
    ws_prior.cell(2, 4, "חובה")
    ws_prior.cell(2, 5, "זכות")
    ws_prior.cell(2, 6, "פ.נ")
    ws_prior.cell(2, 7, "יתרה")
    ws_prior.cell(2, 8, "הערות")
    # Account 1001: debit, balance = 1500
    ws_prior.cell(3, 1, 100)
    ws_prior.cell(3, 2, "1001")
    ws_prior.cell(3, 3, "Cash")
    ws_prior.cell(3, 4, 1000)
    ws_prior.cell(3, 6, 500)
    ws_prior.cell(3, 7, 1500)
    # Account 1002: credit, balance = 600
    ws_prior.cell(4, 1, 100)
    ws_prior.cell(4, 2, "1002")
    ws_prior.cell(4, 3, "Bank")
    ws_prior.cell(4, 5, 200)
    ws_prior.cell(4, 6, 800)
    ws_prior.cell(4, 7, 600)

    # --- New year trial balance ---
    ws_tb = wb.create_sheet(f"מאזן {year}")
    ws_tb.cell(1, 1, f"Company {year}")
    ws_tb.cell(2, 1, "מאזן בוחן")
    ws_tb.cell(3, 1, f"לשנת {year}")
    ws_tb.cell(4, 2, "מיון")
    ws_tb.cell(4, 3, "חשבון")
    ws_tb.cell(4, 4, "שם חשבון")
    ws_tb.cell(4, 5, "חובה")
    ws_tb.cell(4, 6, "זכות")
    ws_tb.cell(4, 7, "הפרש")
    # Section header
    ws_tb.cell(5, 1, "נכסים שוטפים")
    # Account 1001: debit=1200
    ws_tb.cell(6, 2, 100)
    ws_tb.cell(6, 3, "1001")
    ws_tb.cell(6, 4, "Cash")
    ws_tb.cell(6, 5, 1200)
    # Account 1002: credit=250
    ws_tb.cell(7, 2, 100)
    ws_tb.cell(7, 3, "1002")
    ws_tb.cell(7, 4, "Bank")
    ws_tb.cell(7, 6, 250)
    # Group total
    ws_tb.cell(8, 1, 'סה"כ לקבוצה:')
    ws_tb.cell(8, 2, 100)
    ws_tb.cell(8, 5, 1200)
    ws_tb.cell(8, 6, 250)
    ws_tb.cell(8, 7, 950)

    # --- Financial statements sheet (should be preserved) ---
    ws_fin = wb.create_sheet("דוחות כספיים")
    ws_fin.cell(1, 1, "Financial data sentinel")

    path = tmp_path / f"test_workbook_{year}.xlsx"
    wb.save(path)
    return path


def _make_rows(year: int = 2024) -> list[WorkingPaperRow]:
    """Build a minimal set of WorkingPaperRow objects for testing."""
    return [
        WorkingPaperRow(
            row_type="section_header",
            group=None, account_number=None,
            details="נכסים שוטפים",
            debit=None, credit=None, opening_balance=None, notes=None,
        ),
        WorkingPaperRow(
            row_type="account",
            group="100", account_number="1001",
            details="Cash",
            debit=1200.0, credit=None, opening_balance=1500.0, notes=None,
        ),
        WorkingPaperRow(
            row_type="account",
            group="100", account_number="1002",
            details="Bank",
            debit=None, credit=250.0, opening_balance=600.0, notes=None,
        ),
        WorkingPaperRow(
            row_type="group_total",
            group=None, account_number=None,
            details='סה"כ לקבוצה:',
            debit=1200.0, credit=250.0, opening_balance=None, notes=None,
        ),
    ]


def test_sheet_inserted_at_correct_position(tmp_path: Path) -> None:
    """Generated sheet must appear after מאזן <year> and before דוחות כספיים."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    rows = _make_rows()
    output_path = generate_working_paper(data, rows)

    wb_out = openpyxl.load_workbook(output_path)
    names = wb_out.sheetnames
    assert str(data.new_year) in names, f"Sheet '{data.new_year}' not found in {names}"

    new_idx = names.index(str(data.new_year))
    mazn_idx = names.index(f"מאזן {data.new_year}")
    assert new_idx == mazn_idx + 1, (
        f"'{data.new_year}' (idx={new_idx}) should be immediately after "
        f"'מאזן {data.new_year}' (idx={mazn_idx})"
    )

    if "דוחות כספיים" in names:
        fin_idx = names.index("דוחות כספיים")
        assert new_idx < fin_idx, (
            f"'{data.new_year}' (idx={new_idx}) should be before 'דוחות כספיים' (idx={fin_idx})"
        )


def test_sheet_name_is_exact_year_string(tmp_path: Path) -> None:
    """Sheet name must be exactly str(year), nothing more."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    output_path = generate_working_paper(data, _make_rows())

    wb_out = openpyxl.load_workbook(output_path)
    assert str(data.new_year) in wb_out.sheetnames
    # Make sure there's no padded/variant name
    year_str = str(data.new_year)
    for name in wb_out.sheetnames:
        if name.strip() == year_str and name != year_str:
            pytest.fail(f"Found near-match sheet name '{name}' instead of exact '{year_str}'")


def test_no_data_corrupted_in_other_sheets(tmp_path: Path) -> None:
    """Other sheets must be unchanged after workbook generation."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)

    # Capture original values from sheets that should not change
    wb_orig = openpyxl.load_workbook(wb_path, data_only=True)
    orig_fin = list(wb_orig["דוחות כספיים"].iter_rows(values_only=True))
    orig_tb = list(wb_orig[f"מאזן {data.new_year}"].iter_rows(values_only=True))

    output_path = generate_working_paper(data, _make_rows())
    wb_out = openpyxl.load_workbook(output_path, data_only=True)

    out_fin = list(wb_out["דוחות כספיים"].iter_rows(values_only=True))
    out_tb = list(wb_out[f"מאזן {data.new_year}"].iter_rows(values_only=True))

    assert orig_fin == out_fin, "דוחות כספיים sheet was modified"
    assert orig_tb == out_tb, f"מאזן {data.new_year} sheet was modified"


def test_output_xlsx_opens_without_error(tmp_path: Path) -> None:
    """Output .xlsx must be a valid workbook that opens without errors."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    output_path = generate_working_paper(data, _make_rows())

    assert output_path.exists(), f"Output file not created at {output_path}"
    assert output_path.stat().st_size > 0, "Output file is empty"
    # Must open without raising
    wb_check = openpyxl.load_workbook(output_path)
    assert str(data.new_year) in wb_check.sheetnames


def test_new_account_not_in_prior_year_gets_zero_opening_balance(tmp_path: Path) -> None:
    """Account absent from prior year must have opening_balance = 0 in col F."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    # Add a brand-new account (not in prior year)
    rows = _make_rows() + [
        WorkingPaperRow(
            row_type="account",
            group="200", account_number="9999",
            details="New Account",
            debit=500.0, credit=None, opening_balance=0.0, notes=None,
        )
    ]
    output_path = generate_working_paper(data, rows)

    wb_out = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb_out[str(data.new_year)]
    found = False
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) > 1 and str(row[1]) == "9999":
            found = True
            col_f = row[5] if len(row) > 5 else None
            assert col_f == 0.0 or col_f is None or col_f == 0, (
                f"New account 9999: expected F=0, got {col_f}"
            )
    assert found, "Account 9999 not found in generated sheet"


def test_account_in_prior_not_in_new_year_is_absent(tmp_path: Path) -> None:
    """Account from prior year that is absent in new trial balance should not appear."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    # Only include account 1001, not 1002 (which is in prior year)
    rows = [
        WorkingPaperRow(
            row_type="account",
            group="100", account_number="1001",
            details="Cash",
            debit=1200.0, credit=None, opening_balance=1500.0, notes=None,
        ),
    ]
    output_path = generate_working_paper(data, rows)

    wb_out = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb_out[str(data.new_year)]
    for row in ws.iter_rows(min_row=3, values_only=True):
        col_b = str(row[1]) if len(row) > 1 and row[1] is not None else None
        assert col_b != "1002", "Account 1002 (prior-only) should not appear in output"


def test_prior_year_balances_extracted_correctly(tmp_path: Path) -> None:
    """Processor must correctly extract prior year G (יתרה) as opening balances."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    assert "1001" in data.prior_year_balances, "Account 1001 not in prior_year_balances"
    assert "1002" in data.prior_year_balances, "Account 1002 not in prior_year_balances"
    assert data.prior_year_balances["1001"] == 1500.0, (
        f"Expected 1500.0 for 1001, got {data.prior_year_balances['1001']}"
    )
    assert data.prior_year_balances["1002"] == 600.0, (
        f"Expected 600.0 for 1002, got {data.prior_year_balances['1002']}"
    )


def test_d_e_mutual_exclusivity_enforced_by_constructor(tmp_path: Path) -> None:
    """Constructor must raise ValueError when both D and E are set on an account row."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    bad_rows = [
        WorkingPaperRow(
            row_type="account",
            group="100", account_number="1001",
            details="Bad",
            debit=100.0, credit=50.0,  # both set — should raise
            opening_balance=0.0, notes=None,
        )
    ]
    with pytest.raises(ValueError, match="mutual exclusivity"):
        generate_working_paper(data, bad_rows)


def test_negative_debit_raises(tmp_path: Path) -> None:
    """Constructor must raise ValueError for negative debit values."""
    wb_path = make_sample_workbook(tmp_path)
    data = load_workbook_data(wb_path)
    bad_rows = [
        WorkingPaperRow(
            row_type="account",
            group="100", account_number="1001",
            details="Bad",
            debit=-100.0, credit=None,
            opening_balance=0.0, notes=None,
        )
    ]
    with pytest.raises(ValueError):
        generate_working_paper(data, bad_rows)
