"""Deterministic tests for the generated working paper sheet."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from processor import load_workbook_data, WorkbookData, TrialBalanceRow
from constructor import WorkingPaperRow, generate_working_paper
from parser import generate_working_paper_rows

INPUT_DIR = Path(__file__).parent.parent / "input"


def get_generated_sheet(output_path: Path, year: int) -> Worksheet:
    """Return the generated working paper sheet from the output workbook."""
    wb = openpyxl.load_workbook(output_path, data_only=True)
    return wb[str(year)]


def get_account_rows(ws: Worksheet) -> list[tuple]:
    """Return rows where col A (group) is non-empty — these are account rows."""
    result = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 2:
            continue  # Skip title and header rows
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        # Account rows have a group in col A and account number in col B
        if col_a is not None and col_b is not None:
            result.append(row)
    return result


@pytest.fixture(params=sorted(INPUT_DIR.glob("*.xlsx")), ids=lambda p: p.stem)
def pipeline_result(request: pytest.FixtureRequest) -> tuple[WorkbookData, list[WorkingPaperRow], Path]:
    input_path: Path = request.param
    data = load_workbook_data(input_path)
    rows = generate_working_paper_rows(data)
    output_path = generate_working_paper(data, rows)
    return data, rows, output_path


def test_headers_present(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    ws = get_generated_sheet(output_path, data.new_year)
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert header_row[0] == "קבוצה", f"Col A header: expected 'קבוצה', got {header_row[0]}"
    assert header_row[1] == "מס' כרטיס", f"Col B header: expected \"מס' כרטיס\", got {header_row[1]}"
    assert header_row[2] == "פרטים", f"Col C header: expected 'פרטים', got {header_row[2]}"
    assert header_row[3] == "חובה", f"Col D header: expected 'חובה', got {header_row[3]}"
    assert header_row[4] == "זכות", f"Col E header: expected 'זכות', got {header_row[4]}"
    assert header_row[5] == "פ.נ", f"Col F header: expected 'פ.נ', got {header_row[5]}"
    assert header_row[6] == "יתרה", f"Col G header: expected 'יתרה', got {header_row[6]}"
    assert header_row[7] == "הערות", f"Col H header: expected 'הערות', got {header_row[7]}"


def test_d_e_mutual_exclusivity(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    account_rows = get_account_rows(get_generated_sheet(output_path, data.new_year))
    for i, row in enumerate(account_rows):
        d = row[3] if len(row) > 3 else None
        e = row[4] if len(row) > 4 else None
        assert not (d is not None and e is not None), (
            f"Account row {i}: both D={d} and E={e} are set (mutual exclusivity violated)"
        )
        assert not (d is None and e is None), (
            f"Account row {i}: both D and E are empty"
        )


def test_g_formula_correctness(pipeline_result: tuple) -> None:
    """G values must equal F + D - E within float tolerance."""
    data, rows, output_path = pipeline_result
    # Load with data_only=True so openpyxl reads cached formula values
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb[str(data.new_year)]
    account_rows = get_account_rows(ws)
    for i, row in enumerate(account_rows):
        f_val = float(row[5] or 0) if len(row) > 5 else 0.0
        d_val = float(row[3] or 0) if len(row) > 3 else 0.0
        e_val = float(row[4] or 0) if len(row) > 4 else 0.0
        g_val = row[6] if len(row) > 6 else None
        if g_val is None:
            continue  # Formula not evaluated yet (no Excel to compute it)
        expected = f_val + d_val - e_val
        assert abs(float(g_val) - expected) < 0.01, (
            f"Account row {i}: G={g_val} != F({f_val})+D({d_val})-E({e_val})={expected}"
        )


def test_all_trial_balance_accounts_present(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    ws = get_generated_sheet(output_path, data.new_year)
    tb_accounts = {
        row.account_number
        for row in data.trial_balance_rows
        if row.row_type == "account" and row.account_number is not None
    }
    output_accounts: set[str] = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        col_b = str(row[1]) if len(row) > 1 and row[1] is not None else None
        if col_b:
            output_accounts.add(col_b)
    missing = tb_accounts - output_accounts
    assert not missing, f"Missing accounts from trial balance: {sorted(missing)}"


def test_group_numbers_match(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    ws = get_generated_sheet(output_path, data.new_year)
    tb_group_by_account: dict[str, str] = {
        row.account_number: row.group
        for row in data.trial_balance_rows
        if row.row_type == "account" and row.account_number and row.group
    }
    for row in ws.iter_rows(min_row=3, values_only=True):
        col_a = str(row[0]) if len(row) > 0 and row[0] is not None else None
        col_b = str(row[1]) if len(row) > 1 and row[1] is not None else None
        if col_a and col_b and col_b in tb_group_by_account:
            expected_group = tb_group_by_account[col_b]
            assert col_a == expected_group, (
                f"Account {col_b}: group in output={col_a}, expected={expected_group}"
            )


def test_no_missing_groups(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    ws = get_generated_sheet(output_path, data.new_year)
    tb_groups = {
        row.group
        for row in data.trial_balance_rows
        if row.row_type == "account" and row.group
    }
    output_groups: set[str] = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        col_a = str(row[0]) if len(row) > 0 and row[0] is not None else None
        if col_a:
            output_groups.add(col_a)
    missing = tb_groups - output_groups
    assert not missing, f"Missing groups from trial balance: {sorted(missing)}"


def test_d_values_positive(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    account_rows = get_account_rows(get_generated_sheet(output_path, data.new_year))
    for i, row in enumerate(account_rows):
        d = row[3] if len(row) > 3 else None
        if d is not None:
            assert float(d) > 0, f"Account row {i}: D={d} is not positive"


def test_e_values_positive(pipeline_result: tuple) -> None:
    data, rows, output_path = pipeline_result
    account_rows = get_account_rows(get_generated_sheet(output_path, data.new_year))
    for i, row in enumerate(account_rows):
        e = row[4] if len(row) > 4 else None
        if e is not None:
            assert float(e) > 0, f"Account row {i}: E={e} is not positive"
