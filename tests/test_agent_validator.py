"""AI-powered validation tests using Claude CLI as a second-pass validator."""
from __future__ import annotations

import json
import shutil
import subprocess
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from processor import load_workbook_data, WorkbookData
from parser import generate_working_paper_rows
from constructor import generate_working_paper

INPUT_DIR = Path(__file__).parent.parent / "input"
PRD_PATH = Path(__file__).parent.parent / "prd.md"


def serialize_sheet_to_json(ws: Worksheet) -> str:
    """Serialize a worksheet to JSON with column labels."""
    col_labels = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    result = []
    for row in ws.iter_rows(values_only=True):
        row_dict: dict[str, Any] = {}
        for i, val in enumerate(row):
            if i < len(col_labels) and val is not None:
                row_dict[col_labels[i]] = val
        if row_dict:
            result.append(row_dict)
    return json.dumps(result, ensure_ascii=False)


def build_validator_prompt(
    generated_json: str,
    source_json: str,
    prior_json: str,
    new_year: int,
    prior_year: int,
) -> str:
    prd_text = PRD_PATH.read_text(encoding="utf-8") if PRD_PATH.exists() else ""
    return f"""You are a strict validator for a Hebrew financial workbook generator.

## Rules (from prd.md):
{prd_text}

## Task:
Validate the generated {new_year} working paper sheet against the rules above.

## Generated {new_year} working paper (columns A-K):
{generated_json}

## Source מאזן {new_year} trial balance:
{source_json}

## Prior year {prior_year} working paper (columns A-K):
{prior_json}

## Check all of the following:
1. Section header rows precede their respective groups (same order as prior year).
2. Account names (col C) match the trial balance account names.
3. Col F (opening balance) correctly reflects the prior year col G (יתרה) for matching accounts.
4. Group-total rows (סה"כ לקבוצה:) are present and placed correctly.
5. No phantom rows exist (rows in output with no matching trial balance account).
6. Overall sheet structure mirrors the prior year working paper pattern.
7. D and E are mutually exclusive per account row.
8. All account numbers from the trial balance appear in the output.

Return a JSON object with this exact structure:
{{
  "passed": true or false,
  "violations": [
    {{"row": <row_number>, "column": "<col_letter>", "issue": "<description>"}}
  ],
  "summary": "<plain language summary>"
}}

Return ONLY the JSON object, no markdown fences, no explanation."""


def invoke_validator_agent(prompt: str) -> dict[str, Any]:
    """Call Claude CLI as a validator and return the parsed verdict."""
    if shutil.which("claude") is None:
        pytest.skip("claude CLI not found — skipping AI validator test")

    result = subprocess.run(
        ["claude", "-p", prompt],
        capture_output=True,
        text=True,
        timeout=180,
    )
    if result.returncode != 0:
        pytest.fail(f"Claude CLI failed: {result.stderr.strip()}")

    response = result.stdout.strip()
    # Strip markdown fences if present
    if response.startswith("```"):
        lines = response.splitlines()
        inner = []
        in_block = False
        for line in lines:
            if line.startswith("```") and not in_block:
                in_block = True
                continue
            if line.startswith("```") and in_block:
                break
            if in_block:
                inner.append(line)
        response = "\n".join(inner).strip()

    return json.loads(response)


@pytest.fixture(params=sorted(INPUT_DIR.glob("*.xlsx")), ids=lambda p: p.stem)
def validated_pipeline(request: pytest.FixtureRequest) -> tuple[WorkbookData, Path]:
    input_path: Path = request.param
    data = load_workbook_data(input_path)
    rows = generate_working_paper_rows(data)
    output_path = generate_working_paper(data, rows)
    return data, output_path


@pytest.mark.slow
def test_agent_validates_generated_sheet(validated_pipeline: tuple) -> None:
    data, output_path = validated_pipeline

    wb_out = openpyxl.load_workbook(output_path, data_only=True)
    wb_src = openpyxl.load_workbook(data.workbook_path, data_only=True)

    generated_ws = wb_out[str(data.new_year)]
    source_ws = wb_src[f"מאזן {data.new_year}"]
    prior_ws = wb_src[str(data.prior_year)]

    generated_json = serialize_sheet_to_json(generated_ws)
    source_json = serialize_sheet_to_json(source_ws)
    prior_json = serialize_sheet_to_json(prior_ws)

    prompt = build_validator_prompt(
        generated_json, source_json, prior_json, data.new_year, data.prior_year
    )
    verdict = invoke_validator_agent(prompt)

    if not verdict.get("passed") or verdict.get("violations"):
        pytest.fail(
            f"AI validator failed.\nSummary: {verdict.get('summary')}\n"
            f"Violations: {json.dumps(verdict.get('violations', []), ensure_ascii=False, indent=2)}"
        )
