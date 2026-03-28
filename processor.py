"""Processor module: reads .xlsx workbooks and extracts trial balance and prior year data."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Literal, Any
from pathlib import Path
import re

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class TrialBalanceRow:
    row_type: Literal["section_header", "account", "group_total", "title", "header"]
    label: str | None          # col A
    group: str | None          # col B (מיון)
    account_number: str | None # col C (חשבון)
    account_name: str | None   # col D (שם חשבון)
    debit: float | None        # col E (חובה)
    credit: float | None       # col F (זכות)
    net: float | None          # col G (הפרש)


@dataclass
class WorkbookData:
    workbook_path: Path
    new_year: int
    prior_year: int
    trial_balance_rows: list[TrialBalanceRow]
    prior_year_balances: dict[str, float]
    prior_year_sheet_rows: list[list[Any]]


GROUP_TOTAL_MARKER = 'סה"כ לקבוצה:'


def open_workbook(path: Path) -> openpyxl.Workbook:
    """Opens the workbook with data_only=True."""
    return openpyxl.load_workbook(path, data_only=True)


def detect_sheet_pairs(wb: openpyxl.Workbook) -> dict[int, str | None]:
    """Scan sheetnames for pattern מאזן (\\d{4}) and return {year: working_paper_sheet_name_or_None}."""
    pattern = re.compile(r"מאזן (\d{4})")
    sheet_names = set(wb.sheetnames)
    pairs: dict[int, str | None] = {}

    for name in wb.sheetnames:
        match = pattern.fullmatch(name)
        if match:
            year = int(match.group(1))
            working_paper = str(year) if str(year) in sheet_names else None
            pairs[year] = working_paper

    return pairs


def find_target_year(pairs: dict[int, str | None]) -> int:
    """Return the largest year key whose value is None."""
    unpaired = [year for year, wp in pairs.items() if wp is None]
    if not unpaired:
        raise ValueError("No unpaired מאזן year found")
    return max(unpaired)


def find_prior_year(pairs: dict[int, str | None], target_year: int) -> int:
    """Return target_year - 1 if it exists in pairs with a non-None working paper."""
    prior = target_year - 1
    if prior not in pairs:
        raise ValueError(
            f"Prior year {prior} has no matching מאזן sheet in the workbook"
        )
    if pairs[prior] is None:
        raise ValueError(
            f"Prior year {prior} has a מאזן sheet but no working paper sheet ({prior})"
        )
    return prior


def classify_row(row_index: int, row_values: tuple[Any, ...]) -> str | None:
    """Classify a 1-based row into a row type string, or None to skip."""
    if row_index <= 3:
        return "title"
    if row_index == 4:
        return "header"

    # Pad row_values to at least 7 elements
    vals = list(row_values) + [None] * max(0, 7 - len(row_values))

    col_a = vals[0]
    cols_b_to_g = vals[1:7]

    # Completely empty rows
    if all(v is None for v in vals[:7]):
        return None

    # Group total rows
    if col_a == GROUP_TOTAL_MARKER:
        return "group_total"

    # Section header: col A is non-None, cols B–G are all None
    if col_a is not None and all(v is None for v in cols_b_to_g):
        return "section_header"

    # Account row: col A is None, cols B and C are non-None
    col_b = vals[1]
    col_c = vals[2]
    if col_a is None and col_b is not None and col_c is not None:
        return "account"

    return None


def _to_float(value: Any) -> float | None:
    """Cast int or float to float, otherwise return None."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def extract_trial_balance_rows(ws: Worksheet) -> list[TrialBalanceRow]:
    """Extract and classify all rows from a trial balance worksheet."""
    rows: list[TrialBalanceRow] = []

    for row_index, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        # Pad to at least 7 elements
        vals = list(row_values) + [None] * max(0, 7 - len(row_values))

        row_type = classify_row(row_index, tuple(vals))
        if row_type is None:
            continue

        col_a = vals[0]
        col_b = vals[1]
        col_c = vals[2]
        col_d = vals[3]
        col_e = vals[4]
        col_f = vals[5]
        col_g = vals[6]

        label = str(col_a) if col_a is not None else None
        group = str(col_b) if col_b is not None else None
        account_number = str(col_c) if col_c is not None else None
        account_name = str(col_d) if col_d is not None else None
        debit = _to_float(col_e)
        credit = _to_float(col_f)
        net = _to_float(col_g)

        rows.append(TrialBalanceRow(
            row_type=row_type,  # type: ignore[arg-type]
            label=label,
            group=group,
            account_number=account_number,
            account_name=account_name,
            debit=debit,
            credit=credit,
            net=net,
        ))

    return rows


def extract_prior_year_balances(ws: Worksheet) -> dict[str, float]:
    """Extract {account_number: balance} from prior year working paper sheet."""
    balances: dict[str, float] = {}

    for row_index, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip title row (row 1) and header row (row 2); data starts at row 3
        if row_index <= 2:
            continue

        vals = list(row_values) + [None] * max(0, 7 - len(row_values))

        col_b = vals[1]  # account number
        col_g = vals[6]  # יתרה (balance)

        if col_b is None or col_g is None:
            continue

        account_number = str(col_b)
        balance = _to_float(col_g)
        if balance is not None:
            balances[account_number] = balance

    return balances


def extract_prior_year_sheet_rows(ws: Worksheet) -> list[list[Any]]:
    """Return all rows as plain lists."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def load_workbook_data(path: Path) -> WorkbookData:
    """Load and extract all relevant data from the workbook at the given path."""
    wb = open_workbook(path)

    pairs = detect_sheet_pairs(wb)
    new_year = find_target_year(pairs)
    prior_year = find_prior_year(pairs, new_year)

    trial_balance_ws: Worksheet = wb[f"מאזן {new_year}"]

    prior_year_sheet_name = pairs[prior_year] or str(prior_year)
    prior_year_ws: Worksheet = wb[prior_year_sheet_name]

    trial_balance_rows = extract_trial_balance_rows(trial_balance_ws)
    prior_year_balances = extract_prior_year_balances(prior_year_ws)
    prior_year_sheet_rows = extract_prior_year_sheet_rows(prior_year_ws)

    return WorkbookData(
        workbook_path=path,
        new_year=new_year,
        prior_year=prior_year,
        trial_balance_rows=trial_balance_rows,
        prior_year_balances=prior_year_balances,
        prior_year_sheet_rows=prior_year_sheet_rows,
    )
